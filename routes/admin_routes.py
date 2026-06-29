# =========================================================
# IMPORTACIONES
# =========================================================

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    jsonify,
)

from extensions import mongo, socketio
from utils.auth import login_required, role_required

from utils.fraccionamientos import (
    agg_visitas,
    find_visitas,
    contar_visitas,
    find_residentes,
    contar_residentes,
    coleccion_residentes,
    residentes_colecciones,
    buscar_residente_por_id,
    es_fraccionamiento_valido,
    VISITAS_COLECCIONES,
    FRACCIONAMIENTOS,
)

from bson.objectid import ObjectId

from io import BytesIO
from datetime import datetime, timedelta
import os
import time
from werkzeug.security import generate_password_hash

import secrets
import string

from utils.fraccionamientos import obtener_fraccionamientos

# =========================================================
# IMPORT EXCEL
# =========================================================

import pandas as pd
from openpyxl import load_workbook, Workbook

# =========================================================
# REPORTLAB PDF
# =========================================================

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas as _canvas

from flask import jsonify

# =========================================================
# ZONA HORARIA DEL NEGOCIO (Centro de México, UTC-6)
# =========================================================
from datetime import timezone

TZ_LOCAL = timezone(timedelta(hours=-6))


def _ahora_local():
    """Hora actual en México, como datetime naive."""
    return datetime.now(TZ_LOCAL).replace(tzinfo=None)


# =========================================================
# BLUEPRINT
# =========================================================

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# =========================================================
# CACHÉ DE FRACCIONAMIENTOS EN MEMORIA
# Se invalida al crear o eliminar un fraccionamiento.
# Evita consultar MongoDB en cada request del dashboard.
# =========================================================

_CACHE_FRACS = {"lista": None, "ts": 0}
_CACHE_TTL = 60  # segundos


def _fraccionamientos_disponibles():
    """Los definidos en código + los creados en BD. Con caché de 60s."""
    ahora = time.time()
    if _CACHE_FRACS["lista"] is not None and (ahora - _CACHE_FRACS["ts"]) < _CACHE_TTL:
        return _CACHE_FRACS["lista"]

    nombres = list(FRACCIONAMIENTOS)
    for d in mongo.db.fraccionamientos.find({}, {"nombre": 1}):
        nom = (d.get("nombre") or "").strip()
        if nom and nom not in nombres:
            nombres.append(nom)
    resultado = sorted(nombres, key=lambda x: x.lower())

    _CACHE_FRACS["lista"] = resultado
    _CACHE_FRACS["ts"] = ahora
    return resultado


def _invalidar_cache_fracs():
    """Llama esto al crear o eliminar un fraccionamiento."""
    _CACHE_FRACS["lista"] = None
    _CACHE_FRACS["ts"] = 0


# =========================================================
# CACHÉ DE CONTEOS DE RESIDENTES
# =========================================================

_CACHE_CONTEOS = {"data": None, "ts": 0}
_CACHE_CONTEOS_TTL = 120


def _invalidar_cache_conteos():
    """Llama esto al crear/eliminar fraccionamientos o residentes."""
    _CACHE_CONTEOS["data"] = None
    _CACHE_CONTEOS["ts"] = 0


def _residentes_colecciones(db):
    """Devuelve {slug: nombre_coleccion} para cada colección residentes_*."""
    return {
        nombre[len("residentes_") :]: nombre
        for nombre in db.list_collection_names()
        if nombre.startswith("residentes_")
    }


def _conteos_residentes_cacheados(lista_fracs):
    ahora = time.time()

    if (
        _CACHE_CONTEOS["data"] is not None
        and (ahora - _CACHE_CONTEOS["ts"]) < _CACHE_CONTEOS_TTL
    ):
        return _CACHE_CONTEOS["data"]

    total_global = 0
    conteos = {f: 0 for f in lista_fracs}

    for col_name in _residentes_colecciones(mongo.db).values():
        pipeline = [
            {"$match": {"rol": "residente"}},
            {"$group": {"_id": "$fraccionamiento", "n": {"$sum": 1}}},
        ]

        for row in mongo.db[col_name].aggregate(pipeline):
            frac_val = (row.get("_id") or "").strip().lower()
            n = row.get("n", 0)
            total_global += n

            for f in lista_fracs:
                if f.strip().lower() == frac_val:
                    conteos[f] += n
                    break

    resultado = (conteos, total_global)
    _CACHE_CONTEOS["data"] = resultado
    _CACHE_CONTEOS["ts"] = ahora
    return resultado


# =========================================================
# SISTEMA DE DISEÑO DE REPORTES PDF
# =========================================================

BRAND_DARK = colors.HexColor("#0F172A")
BRAND_ACCENT = colors.HexColor("#2563EB")
BRAND_GRID = colors.HexColor("#E2E8F0")
TEXT_DARK = colors.HexColor("#1E293B")
TEXT_MUTED = colors.HexColor("#64748B")
ROW_ALT = colors.HexColor("#F8FAFC")
OK_GREEN = colors.HexColor("#16A34A")
WARN_RED = colors.HexColor("#DC2626")
AMBER = colors.HexColor("#D97706")
CYAN = colors.HexColor("#0891B2")

ESTADO_COLORES = {
    "activo": OK_GREEN,
    "permitido": OK_GREEN,
    "autorizado": OK_GREEN,
    "resuelta": OK_GREEN,
    "resuelto": OK_GREEN,
    "dentro": BRAND_ACCENT,
    "salida_registrada": CYAN,
    "finalizada": CYAN,
    "inactivo": TEXT_MUTED,
    "bloqueado": TEXT_MUTED,
    "cerrada": TEXT_MUTED,
    "cancelado": WARN_RED,
    "rechazado": WARN_RED,
    "vencido": AMBER,
    "pendiente": AMBER,
    "pendiente_autorizacion": AMBER,
    "abierta": AMBER,
}

_CELDA = ParagraphStyle(
    "celda",
    fontName="Helvetica",
    fontSize=8,
    leading=10,
    textColor=TEXT_DARK,
    wordWrap="CJK",
)
_CELDA_NUM = ParagraphStyle(
    "celdanum", parent=_CELDA, alignment=1, textColor=TEXT_MUTED
)
_CELDA_FUERTE = ParagraphStyle(
    "celdaf", parent=_CELDA, fontName="Helvetica-Bold", textColor=BRAND_DARK
)


def _c(valor, estilo=_CELDA):
    return Paragraph("" if valor is None else str(valor), estilo)


def _badge(texto, fondo, color_texto=colors.white):
    est = ParagraphStyle(
        "badge",
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=color_texto,
        alignment=1,
        backColor=fondo,
        borderPadding=(2, 6, 2, 6),
        borderRadius=6,
        leading=10,
    )
    return Paragraph((texto or "—").upper(), est)


def _badge_estado(estado_raw, etiqueta=None):
    color = ESTADO_COLORES.get((estado_raw or "").lower(), TEXT_MUTED)
    return _badge(etiqueta or estado_raw or "—", color)


def _kpi_card(valor, etiqueta, color=BRAND_ACCENT):
    ev = ParagraphStyle(
        "kv",
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=BRAND_DARK,
        alignment=0,
        leading=22,
    )
    el = ParagraphStyle(
        "kl",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        textColor=TEXT_MUTED,
        alignment=0,
        leading=11,
        spaceBefore=3,
    )
    inner = Table([[_c(valor, ev)], [_c(etiqueta.upper(), el)]])
    inner.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.8, BRAND_GRID),
                ("LINEBEFORE", (0, 0), (0, -1), 3.5, color),
                ("TOPPADDING", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 0),
                ("TOPPADDING", (0, 1), (-1, 1), 0),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 11),
                ("LEFTPADDING", (0, 0), (-1, -1), 13),
                ("RIGHTPADDING", (0, 0), (-1, -1), 13),
                ("ROUNDEDCORNERS", [7, 7, 7, 7]),
            ]
        )
    )
    return inner


def _fila_kpis(cards, ancho_total):
    n = len(cards)
    gap = 12
    cw = (ancho_total - gap * (n - 1)) / n
    row, widths = [], []
    for i, card in enumerate(cards):
        row.append(card)
        widths.append(cw)
        if i < n - 1:
            row.append("")
            widths.append(gap)
    t = Table([row], colWidths=widths)
    t.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return t


def _tabla_datos(data, col_widths, aligns=None):
    t = Table(data, repeatRows=1, colWidths=col_widths)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("LINEBELOW", (0, 1), (-1, -1), 0.4, BRAND_GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 1), (-1, -1), "LEFT"),
        ("TOPPADDING", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("ROUNDEDCORNERS", [6, 6, 0, 0]),
    ]
    if aligns:
        for col, al in enumerate(aligns):
            estilo.append(("ALIGN", (col, 1), (col, -1), al))
    t.setStyle(TableStyle(estilo))
    return t


def _make_canvas_class(titulo, subtitulo, pagesize):
    class _ReporteCanvas(_canvas.Canvas):
        def __init__(self, *a, **k):
            super().__init__(*a, **k)
            self._guardadas = []

        def showPage(self):
            self._guardadas.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._guardadas)
            for i, st in enumerate(self._guardadas, start=1):
                self.__dict__.update(st)
                self._header()
                self._footer(i, total)
                super().showPage()
            super().save()

        def _header(self):
            w, h = pagesize
            bh = 62
            self.setFillColor(BRAND_DARK)
            self.rect(0, h - bh, w, bh, fill=1, stroke=0)
            self.setFillColor(BRAND_ACCENT)
            self.rect(0, h - bh - 4, w, 4, fill=1, stroke=0)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 17)
            self.drawString(40, h - 33, titulo)
            if subtitulo:
                self.setFillColor(colors.HexColor("#CBD5E1"))
                self.setFont("Helvetica", 9)
                self.drawString(40, h - 49, subtitulo)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 13)
            self.drawRightString(w - 40, h - 30, "AccessQR")
            self.setFillColor(colors.HexColor("#94A3B8"))
            self.setFont("Helvetica", 8)
            self.drawRightString(w - 40, h - 44, "Sistema Residencial")

        def _footer(self, pagina, total):
            w, h = pagesize
            self.setStrokeColor(BRAND_GRID)
            self.setLineWidth(0.7)
            self.line(40, 40, w - 40, 40)
            self.setFillColor(TEXT_MUTED)
            self.setFont("Helvetica", 8)
            self.drawString(40, 28, "AccessQR \u00b7 Sistema Residencial")
            self.drawCentredString(
                w / 2,
                28,
                datetime.now().strftime("Generado el %d/%m/%Y %I:%M %p"),
            )
            self.drawRightString(w - 40, 28, f"P\u00e1gina {pagina} de {total}")

    return _ReporteCanvas


def _construir_reporte_pdf(buffer, titulo, subtitulo, contenido, asunto=None):
    pagesize = landscape(letter)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=40,
        rightMargin=40,
        topMargin=94,
        bottomMargin=54,
        title=f"Acceso QR | {titulo}",
        author="AccessQR",
        subject=asunto or titulo,
        creator="AccessQR | Sistema Residencial",
    )
    doc.build(contenido, canvasmaker=_make_canvas_class(titulo, subtitulo, pagesize))


def _ancho_util():
    return landscape(letter)[0] - 80


def _registrar_historial_reporte(nombre, tipo, formato):
    from flask import session

    mongo.db.reportes.insert_one(
        {
            "nombre": nombre,
            "tipo": tipo,
            "formato": formato,
            "usuario": session.get("nombre", "Administrador"),
            "fecha": datetime.now(),
            "estado": "Generado",
        }
    )


def _distinct_visitas(campo, filtro):
    valores = set()
    for nombre in VISITAS_COLECCIONES.values():
        valores.update(mongo.db[nombre].distinct(campo, filtro))
    return list(valores)


# =========================================================
# ÍNDICES — se crean UNA sola vez por proceso
# =========================================================

_INDICES_LISTOS = False


def _asegurar_indices():
    global _INDICES_LISTOS
    if _INDICES_LISTOS:
        return
    try:
        for nombre in VISITAS_COLECCIONES.values():
            col = mongo.db[nombre]
            col.create_index("fecha_visita")
            col.create_index([("fraccionamiento", 1), ("fecha_visita", 1)])
            col.create_index([("fecha_visita", 1), ("hora_inicio", 1)])
            col.create_index([("fecha_visita", 1), ("created_at", 1)])
            col.create_index("estado")
            col.create_index("residente_nombre")
            col.create_index("nombre_visitante")
            col.create_index("vehiculo.placa")

        mongo.db.access_logs.create_index([("fecha_hora", -1)])
        mongo.db.access_logs.create_index([("resultado", 1), ("fecha_hora", -1)])
        mongo.db.incidencias.create_index([("fecha_hora", -1)])
        mongo.db.reportes.create_index([("fecha", -1)])
        mongo.db.users.create_index([("rol", 1)])

        for nombre in mongo.db.list_collection_names():
            if nombre.startswith("residentes_"):
                # Índices originales
                mongo.db[nombre].create_index([("rol", 1), ("nombre", 1)])
                mongo.db[nombre].create_index("correo")
                mongo.db[nombre].create_index(
                    [("fraccionamiento", 1), ("rol", 1), ("nombre", 1)]
                )
                mongo.db[nombre].create_index([("nombre", 1)])
                mongo.db[nombre].create_index([("rol", 1), ("fraccionamiento", 1)])

        _INDICES_LISTOS = True
    except Exception:
        pass


def _facet_visitas(db, match):
    """Una sola agregación sobre las 3 colecciones con $unionWith."""
    cols = list(VISITAS_COLECCIONES.values())
    if not cols:
        return {}

    base = cols[0]
    pipeline = [{"$match": match}]
    for otra in cols[1:]:
        pipeline.append({"$unionWith": {"coll": otra, "pipeline": [{"$match": match}]}})

    pipeline.append(
        {
            "$facet": {
                "estado": [{"$group": {"_id": "$estado", "n": {"$sum": 1}}}],
                "tipos": [
                    {
                        "$group": {
                            "_id": {"$ifNull": ["$modalidad_visita", "General"]},
                            "n": {"$sum": 1},
                        }
                    },
                    {"$sort": {"n": -1}},
                ],
                "dias": [{"$group": {"_id": "$fecha_visita", "n": {"$sum": 1}}}],
                "dias_salidas": [
                    {"$match": {"estado": "salida_registrada"}},
                    {"$group": {"_id": "$fecha_visita", "n": {"$sum": 1}}},
                ],
                "residentes": [
                    {"$group": {"_id": "$residente_nombre", "n": {"$sum": 1}}},
                    {"$sort": {"n": -1}},
                    {"$limit": 10},
                ],
                "vehiculos": [
                    {"$match": {"vehiculo.placa": {"$nin": [None, ""]}}},
                    {"$group": {"_id": "$vehiculo.placa"}},
                    {"$limit": 40},
                ],
            }
        }
    )

    try:
        res = list(db[base].aggregate(pipeline, allowDiskUse=True))
        return res[0] if res else {}
    except Exception:
        return {}


def _generar_password(longitud=10):
    alfabeto = string.ascii_letters + string.digits
    base = "".join(secrets.choice(alfabeto) for _ in range(longitud))
    return base + "*"


# =========================================================
# ENDPOINT: Última actualización de configuración
# =========================================================


@admin_bp.route("/configuracion/ultima-actualizacion/<fraccionamiento>")
@login_required
@role_required("admin")
def ultima_actualizacion(fraccionamiento):
    config = mongo.db.configuraciones.find_one({"fraccionamiento": fraccionamiento})
    return jsonify(
        {
            "hora": (
                config.get("actualizado_str", "Sin cambios")
                if config
                else "Sin cambios"
            )
        }
    )


# =========================================================
# ENDPOINT AJAX: KPIs y datos de gráficas (carga lazy)
# El dashboard HTML carga instantáneo con ceros y luego
# este endpoint rellena los datos reales en ~500ms.
# =========================================================


@admin_bp.route("/dashboard/kpis")
@login_required
@role_required("admin")
def dashboard_kpis():
    from datetime import datetime as _dt

    modo = request.args.get("modo", "dia")
    frac_sel = request.args.get("fraccionamiento", "").strip().lower()
    dia_sel = request.args.get("dia", "").strip()
    semana_sel = request.args.get("semana", "").strip()
    mes_sel = request.args.get("mes", "").strip()
    hoy = _ahora_local()

    # ── Rango del período ──
    if modo == "semana":
        if not semana_sel:
            iso = hoy.isocalendar()
            semana_sel = f"{iso[0]}-W{iso[1]:02d}"
        anio, sem = semana_sel.split("-W")
        ini = _dt.strptime(f"{anio}-W{int(sem):02d}-1", "%G-W%V-%u")
        fin = ini + timedelta(days=6)
    elif modo == "mes":
        if not mes_sel:
            mes_sel = hoy.strftime("%Y-%m")
        anio, mes = map(int, mes_sel.split("-"))
        ini = datetime(anio, mes, 1)
        fin = (
            datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        ) - timedelta(days=1)
    else:
        if not dia_sel:
            dia_sel = hoy.strftime("%Y-%m-%d")
        ini = _dt.strptime(dia_sel, "%Y-%m-%d")
        fin = ini

    inicio_str = ini.strftime("%Y-%m-%d")
    fin_str = fin.strftime("%Y-%m-%d")
    dias_periodo = max(1, (fin - ini).days + 1)

    match = {"fecha_visita": {"$gte": inicio_str, "$lte": fin_str}}
    if frac_sel:
        match["fraccionamiento"] = frac_sel

    _asegurar_indices()
    facet = _facet_visitas(mongo.db, match)

    estado_counts = {r["_id"]: r["n"] for r in facet.get("estado", []) if r.get("_id")}
    total_visitas = sum(estado_counts.values())
    dentro = estado_counts.get("dentro", 0)
    activas = estado_counts.get("activo", 0)
    salidas = estado_counts.get("salida_registrada", 0)

    dias_raw = {r["_id"]: r["n"] for r in facet.get("dias", []) if r.get("_id")}
    fechas_ord = sorted(dias_raw.keys(), key=lambda x: _dt.strptime(x, "%Y-%m-%d"))
    salidas_raw = {
        r["_id"]: r["n"] for r in facet.get("dias_salidas", []) if r.get("_id")
    }

    tipo_labels, tipo_data = [], []
    for row in facet.get("tipos", []):
        tipo_labels.append(row["_id"] or "General")
        tipo_data.append(row["n"])

    rechazados = mongo.db.access_logs.count_documents({"resultado": "rechazado"})
    total_residentes = contar_residentes(mongo.db, {"rol": "residente"})
    total_guardias = mongo.db.users.count_documents({"rol": "guardia"})
    total_incidencias = mongo.db.incidencias.count_documents({})

    from concurrent.futures import ThreadPoolExecutor

    def _q_rechazados():
        return mongo.db.access_logs.count_documents({"resultado": "rechazado"})

    def _q_residentes():
        return contar_residentes(mongo.db, {"rol": "residente"})

    def _q_guardias():
        return mongo.db.users.count_documents({"rol": "guardia"})

    def _q_incidencias():
        return mongo.db.incidencias.count_documents({})

    with ThreadPoolExecutor(max_workers=4) as ex:
        f_rech = ex.submit(_q_rechazados)
        f_res = ex.submit(_q_residentes)
        f_gua = ex.submit(_q_guardias)
        f_inc = ex.submit(_q_incidencias)
        rechazados = f_rech.result()
        total_residentes = f_res.result()
        total_guardias = f_gua.result()
        total_incidencias = f_inc.result()

    return jsonify(
        {
            "total_visitas": total_visitas,
            "dentro": dentro,
            "activas": activas,
            "salidas": salidas,
            "rechazados": rechazados,
            "total_residentes": total_residentes,
            "total_guardias": total_guardias,
            "total_incidencias": total_incidencias,
            "promedio": round(total_visitas / dias_periodo, 1),
            "tipo_labels": tipo_labels,
            "tipo_data": tipo_data,
            "dias_labels": [
                _dt.strptime(f, "%Y-%m-%d").strftime("%d/%m") for f in fechas_ord
            ],
            "dias_data": [dias_raw[f] for f in fechas_ord],
            "salidas_data": [salidas_raw.get(f, 0) for f in fechas_ord],
            "top_residentes": [
                [r["_id"] or "—", r["n"]] for r in facet.get("residentes", [])[:5]
            ],
            "vehiculos": [r["_id"] for r in facet.get("vehiculos", []) if r.get("_id")][
                :10
            ],
            "actividad_reciente": [],  # se carga por el partial de tabla
        }
    )


# =========================================================
# DASHBOARD ADMIN — solo sirve el HTML, sin agregaciones
# =========================================================


@admin_bp.route("/dashboard")
@login_required
@role_required("admin")
def dashboard():
    from datetime import datetime as _dt

    MESES_ES = [
        "",
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ]

    modo = request.args.get("modo", "dia")
    frac_sel = request.args.get("fraccionamiento", "").strip().lower()
    dia_sel = request.args.get("dia", "").strip()
    semana_sel = request.args.get("semana", "").strip()
    mes_sel = request.args.get("mes", "").strip()
    busqueda = request.args.get("busqueda", "").strip()
    fecha_inicio_tabla = request.args.get("fecha_inicio_tabla", "").strip()
    fecha_fin_tabla = request.args.get("fecha_fin_tabla", "").strip()

    hoy = _ahora_local()

    # ── Solo calcula el label del período (sin tocar MongoDB) ──
    if modo == "semana":
        if not semana_sel:
            iso = hoy.isocalendar()
            semana_sel = f"{iso[0]}-W{iso[1]:02d}"
        anio, sem = semana_sel.split("-W")
        ini = _dt.strptime(f"{anio}-W{int(sem):02d}-1", "%G-W%V-%u")
        fin = ini + timedelta(days=6)
        rango_label = (
            f"Semana del {ini.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
        )
        periodo_qs = f"modo=semana&semana={semana_sel}"
    elif modo == "mes":
        if not mes_sel:
            mes_sel = hoy.strftime("%Y-%m")
        anio, mes = map(int, mes_sel.split("-"))
        rango_label = f"{MESES_ES[mes]} {anio}"
        periodo_qs = f"modo=mes&mes={mes_sel}"
    else:
        modo = "dia"
        if not dia_sel:
            dia_sel = hoy.strftime("%Y-%m-%d")
        d = _dt.strptime(dia_sel, "%Y-%m-%d")
        rango_label = f"{d.day} de {MESES_ES[d.month]} {d.year}"
        periodo_qs = f"modo=dia&dia={dia_sel}"

    # ── Fraccionamientos desde caché ──
    fraccionamientos = _fraccionamientos_disponibles()

    PER_PAGE_OPCIONES = [10, 25, 50, 100]
    por_pagina_tabla = int(request.args.get("per_page", 10))
    if por_pagina_tabla not in PER_PAGE_OPCIONES:
        por_pagina_tabla = 10

    return render_template(
        "admin_dashboard.html",
        modo=modo,
        filtro=modo,
        dia_sel=dia_sel,
        semana_sel=semana_sel,
        mes_sel=mes_sel,
        rango_label=rango_label,
        periodo_qs=periodo_qs,
        frac_sel=frac_sel,
        fraccionamientos=fraccionamientos,
        busqueda=busqueda,
        fecha_inicio_tabla=fecha_inicio_tabla,
        fecha_fin_tabla=fecha_fin_tabla,
        # Todo en 0/vacío — JS lo rellena
        total_visitas=0,
        dentro=0,
        activas=0,
        salidas=0,
        rechazados=0,
        total_residentes=0,
        total_guardias=0,
        total_incidencias=0,
        promedio=0,
        pendientes_autorizacion=0,
        tipo_labels=[],
        tipo_data=[],
        dias_labels=[],
        dias_data=[],
        salidas_data=[],
        top_residentes=[],
        vehiculos=[],
        actividad_reciente=[],
        visitas_tabla=[],
        pagina_tabla=1,
        total_paginas_tabla=1,
        total_registros_tabla=0,
        por_pagina_tabla=por_pagina_tabla,
        per_page_opciones=PER_PAGE_OPCIONES,
        horas_labels=[],
        horas_data=[],
        privadas_labels=[],
        privadas_data=[],
        visitas_dentro=[],
        alertas=[],
        accesos_rechazados=[],
        top_guardias=[],
        residentes_activos=[],
    )


# =========================================================
# BÚSQUEDA / PAGINACIÓN INSTANTÁNEA DE LA TABLA (AJAX)
# =========================================================


@admin_bp.route("/visitas/tabla")
@login_required
@role_required("admin")
def buscar_visitas_tabla():
    import re
    from datetime import datetime as _dt

    modo = request.args.get("modo", "dia")
    busqueda = request.args.get("busqueda", "").strip()
    fecha_inicio_tabla = request.args.get("fecha_inicio_tabla", "").strip()
    fecha_fin_tabla = request.args.get("fecha_fin_tabla", "").strip()
    frac_sel = request.args.get("fraccionamiento", "").strip().lower()
    dia_sel = request.args.get("dia", "").strip()
    semana_sel = request.args.get("semana", "").strip()
    mes_sel = request.args.get("mes", "").strip()

    hoy = _ahora_local()

    if modo == "semana":
        if not semana_sel:
            iso = hoy.isocalendar()
            semana_sel = f"{iso[0]}-W{iso[1]:02d}"
        anio, sem = semana_sel.split("-W")
        ini = _dt.strptime(f"{anio}-W{int(sem):02d}-1", "%G-W%V-%u")
        fin = ini + timedelta(days=6)
        periodo_qs = f"modo=semana&semana={semana_sel}"
    elif modo == "mes":
        if not mes_sel:
            mes_sel = hoy.strftime("%Y-%m")
        anio, mes = map(int, mes_sel.split("-"))
        ini = datetime(anio, mes, 1)
        fin = (
            datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        ) - timedelta(days=1)
        periodo_qs = f"modo=mes&mes={mes_sel}"
    else:
        modo = "dia"
        if not dia_sel:
            dia_sel = hoy.strftime("%Y-%m-%d")
        ini = _dt.strptime(dia_sel, "%Y-%m-%d")
        fin = ini
        periodo_qs = f"modo=dia&dia={dia_sel}"

    inicio_str = ini.strftime("%Y-%m-%d")
    fin_str = fin.strftime("%Y-%m-%d")

    if fecha_inicio_tabla or fecha_fin_tabla:
        inicio_str = fecha_inicio_tabla or inicio_str
        fin_str = fecha_fin_tabla or fin_str
        if inicio_str > fin_str:
            inicio_str, fin_str = fin_str, inicio_str

    match_visitas = {"fecha_visita": {"$gte": inicio_str, "$lte": fin_str}}
    if frac_sel:
        match_visitas["fraccionamiento"] = frac_sel
    if busqueda:
        rx = {"$regex": re.escape(busqueda), "$options": "i"}
        match_visitas["$or"] = [
            {"nombre_visitante": rx},
            {"residente_nombre": rx},
            {"vehiculo.placa": rx},
        ]

    PER_PAGE_OPCIONES = [10, 25, 50, 100]
    pagina_tabla = max(1, int(request.args.get("page", 1)))
    por_pagina_tabla = int(request.args.get("per_page", 10))
    if por_pagina_tabla not in PER_PAGE_OPCIONES:
        por_pagina_tabla = 10

    total = contar_visitas(mongo.db, match_visitas)
    total_paginas = max(1, (total + por_pagina_tabla - 1) // por_pagina_tabla)
    if pagina_tabla > total_paginas:
        pagina_tabla = total_paginas

    visitas = find_visitas(
        mongo.db,
        match_visitas,
        sort=[("fecha_visita", 1), ("created_at", 1)],
        skip=(pagina_tabla - 1) * por_pagina_tabla,
        limit=por_pagina_tabla,
        frac=frac_sel or None,
    )

    return render_template(
        "partials/_visitas_tabla.html",
        visitas_tabla=visitas,
        busqueda=busqueda,
        frac_sel=frac_sel,
        periodo_qs=periodo_qs,
        fecha_inicio_tabla=fecha_inicio_tabla,
        fecha_fin_tabla=fecha_fin_tabla,
        modo=modo,
        dia_sel=dia_sel,
        semana_sel=semana_sel,
        mes_sel=mes_sel,
        pagina_tabla=pagina_tabla,
        total_paginas_tabla=total_paginas,
        total_registros_tabla=total,
        por_pagina_tabla=por_pagina_tabla,
        per_page_opciones=PER_PAGE_OPCIONES,
    )


# =====================================================
# GESTIÓN DE RESIDENTES
# =====================================================


@admin_bp.route("/residentes")
@login_required
@role_required("admin")
def residentes():
    import re

    pagina = int(request.args.get("page", 1))
    por_pagina = 25
    frac_sel = request.args.get("fraccionamiento", "").strip()
    busqueda = request.args.get("busqueda", "").strip()

    filtro = {"rol": "residente"}
    if frac_sel:
        filtro["fraccionamiento"] = {"$regex": f"^{frac_sel}$", "$options": "i"}
    if busqueda:
        rx = {"$regex": re.escape(busqueda), "$options": "i"}
        filtro["$or"] = [
            {"nombre": rx},
            {"correo": rx},
            {"telefono": rx},
            {"privada": rx},
            {"numero_casa": rx},
        ]

    lista_fracs = _fraccionamientos_disponibles()
    conteos, total_global = _conteos_residentes_cacheados(lista_fracs)

    total = conteos.get(frac_sel, 0) if frac_sel else total_global
    if busqueda:
        total = contar_residentes(mongo.db, filtro)

    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    if pagina > total_paginas:
        pagina = total_paginas

    usuarios = find_residentes(
        mongo.db,
        filtro,
        sort=[("nombre", 1)],
        skip=(pagina - 1) * por_pagina,
        limit=por_pagina,
    )

    return render_template(
        "admin_residentes.html",
        usuarios=usuarios,
        pagina=pagina,
        total_paginas=total_paginas,
        fraccionamientos=lista_fracs,
        frac_sel=frac_sel,
        busqueda=busqueda,
        conteos=conteos,
        total=total,
        total_global=total_global,
    )


# =========================================
# REGISTRAR RESIDENTE
# =========================================


@admin_bp.route("/residentes/registrar", methods=["GET", "POST"])
@login_required
@role_required("admin")
def registrar_residente():
    if request.method == "POST":
        nombre = request.form["nombre"].strip()
        correo = request.form["correo"].strip().lower()
        telefono = request.form["telefono"].strip()
        fraccionamiento = request.form["fraccionamiento"].strip().lower()
        privada = request.form["privada"].strip().lower()
        numero_casa = request.form["numero_casa"].strip().lower()

        if not es_fraccionamiento_valido(fraccionamiento):
            flash("Selecciona un fraccionamiento válido.", "danger")
            return redirect(url_for("admin.registrar_residente"))

        residentes_col = coleccion_residentes(mongo.db, fraccionamiento)

        from utils.fraccionamientos import correo_ya_existe

        if correo_ya_existe(mongo.db, correo):
            flash("El correo electrónico ya se encuentra registrado.", "danger")
            return redirect(url_for("admin.registrar_residente"))

        casa_existente = residentes_col.find_one(
            {
                "rol": "residente",
                "fraccionamiento": fraccionamiento,
                "privada": privada,
                "numero_casa": numero_casa,
            }
        )
        if casa_existente:
            flash(
                f"La casa {numero_casa.upper()} ya está registrada en "
                f"{privada.title()} - {fraccionamiento.title()}",
                "danger",
            )
            return redirect(url_for("admin.registrar_residente"))

        residentes_col.insert_one(
            {
                "nombre": nombre,
                "correo": correo,
                "password": generate_password_hash("Residente123*"),
                "telefono": telefono,
                "fraccionamiento": fraccionamiento,
                "privada": privada,
                "numero_casa": numero_casa,
                "estado": "activo",
                "rol": "residente",
                "created_at": datetime.now(),
                "ultimo_acceso": None,
                "intentos_fallidos": 0,
                "bloqueado_hasta": None,
            }
        )

        _invalidar_cache_conteos()

        socketio.emit("actualizar_residentes", to="rol:admin")
        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash("Residente registrado correctamente.", "success")
        return redirect(url_for("admin.residentes"))

    return render_template(
        "registrar_residente.html",
        fraccionamientos=_fraccionamientos_disponibles(),
    )


# =========================================
# CARGA MASIVA DE RESIDENTES (EXCEL)
# =========================================

COLUMNAS_RESIDENTE = [
    "nombre",
    "correo",
    "telefono",
    "fraccionamiento",
    "privada",
    "numero_casa",
]


@admin_bp.route("/residentes/plantilla")
@login_required
@role_required("admin")
def descargar_plantilla_residentes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Residentes"
    ws.append(COLUMNAS_RESIDENTE)
    ws.append(
        [
            "Mariana Ríos López",
            "mariana@ejemplo.com",
            "555-123-4567",
            "El Porvenir",
            "Privada Los Robles",
            "A-14",
        ]
    )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_residentes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/residentes/carga-masiva", methods=["POST"])
@login_required
@role_required("admin")
def carga_masiva_residentes():
    import traceback
    from utils.fraccionamientos import correo_ya_existe

    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename.lower().endswith((".xlsx", ".xls")):
        flash("Sube un archivo Excel válido (.xlsx).", "danger")
        return redirect(url_for("admin.registrar_residente"))

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        flash("No se pudo leer el archivo.", "danger")
        return redirect(url_for("admin.registrar_residente"))

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        flash("El archivo está vacío.", "warning")
        return redirect(url_for("admin.registrar_residente"))

    encabezados = [str(c).strip().lower() if c is not None else "" for c in filas[0]]
    faltantes = [c for c in COLUMNAS_RESIDENTE if c not in encabezados]
    if faltantes:
        flash(f"Faltan columnas: {', '.join(faltantes)}.", "danger")
        return redirect(url_for("admin.registrar_residente"))

    idx = {col: encabezados.index(col) for col in COLUMNAS_RESIDENTE}
    password_temporal = generate_password_hash("Residente123*")
    disponibles_norm = {f.strip().lower() for f in _fraccionamientos_disponibles()}

    creados = 0
    omitidos = 0
    errores = []
    correos_archivo = set()

    for n, fila in enumerate(filas[1:], start=2):
        try:

            def val(col):
                v = fila[idx[col]] if idx[col] < len(fila) else None
                return str(v).strip() if v is not None else ""

            nombre = val("nombre")
            correo = val("correo").lower()
            telefono = val("telefono")
            fraccionamiento = val("fraccionamiento").lower()
            privada = val("privada").lower()
            numero_casa = val("numero_casa").lower()

            if not all(
                [nombre, correo, telefono, fraccionamiento, privada, numero_casa]
            ):
                omitidos += 1
                errores.append(f"Fila {n}: datos incompletos.")
                continue

            if fraccionamiento not in disponibles_norm:
                omitidos += 1
                errores.append(
                    f"Fila {n}: fraccionamiento '{fraccionamiento}' no válido."
                )
                continue

            slug = _slug_fraccionamiento(fraccionamiento)
            col_res = mongo.db[f"residentes_{slug}"]

            if correo in correos_archivo:
                omitidos += 1
                errores.append(f"Fila {n}: correo {correo} repetido en el archivo.")
                continue

            try:
                if correo_ya_existe(mongo.db, correo):
                    omitidos += 1
                    errores.append(f"Fila {n}: correo {correo} ya registrado.")
                    continue
            except Exception:
                if col_res.find_one({"correo": correo}):
                    omitidos += 1
                    errores.append(f"Fila {n}: correo {correo} ya registrado.")
                    continue

            if col_res.find_one(
                {
                    "rol": "residente",
                    "fraccionamiento": fraccionamiento,
                    "privada": privada,
                    "numero_casa": numero_casa,
                }
            ):
                omitidos += 1
                errores.append(f"Fila {n}: casa {numero_casa.upper()} ya existe.")
                continue

            col_res.insert_one(
                {
                    "nombre": nombre,
                    "correo": correo,
                    "password": password_temporal,
                    "telefono": telefono,
                    "fraccionamiento": fraccionamiento,
                    "privada": privada,
                    "numero_casa": numero_casa,
                    "estado": "activo",
                    "rol": "residente",
                    "created_at": datetime.now(),
                    "ultimo_acceso": None,
                    "intentos_fallidos": 0,
                    "bloqueado_hasta": None,
                }
            )
            correos_archivo.add(correo)
            creados += 1

        except Exception as e:
            omitidos += 1
            errores.append(f"Fila {n}: error inesperado ({type(e).__name__}).")
            traceback.print_exc()
            continue

    if creados:
        _invalidar_cache_conteos()

        socketio.emit("actualizar_residentes", to="rol:admin")
        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash(f"{creados} residente(s) registrados correctamente.", "success")
    if omitidos:
        detalle = " ".join(errores[:8])
        extra = f" (+{len(errores) - 8} más)" if len(errores) > 8 else ""
        flash(f"{omitidos} fila(s) omitidas. {detalle}{extra}", "warning")
    if not creados and not omitidos:
        flash("No se procesó ninguna fila.", "warning")

    return redirect(url_for("admin.registrar_residente"))


# =========================================
# VER / EDITAR / BLOQUEAR / ELIMINAR / DESBLOQUEAR RESIDENTE
# =========================================


@admin_bp.route("/residentes/<id>")
@login_required
@role_required("admin")
def ver_residente(id):
    usuario, _ = buscar_residente_por_id(mongo.db, id)
    return render_template("ver_residente.html", usuario=usuario)


@admin_bp.route("/residentes/editar/<id>", methods=["GET", "POST"])
@login_required
@role_required("admin")
def editar_residente(id):
    usuario, col = buscar_residente_por_id(mongo.db, id)
    if request.method == "POST":
        col.update_one(
            {"_id": ObjectId(id)},
            {
                "$set": {
                    "nombre": request.form["nombre"],
                    "correo": request.form["correo"],
                    "telefono": request.form["telefono"],
                    "fraccionamiento": request.form["fraccionamiento"],
                    "privada": request.form["privada"],
                    "numero_casa": request.form["numero_casa"],
                }
            },
        )
        socketio.emit("actualizar_residentes", to="rol:admin")
        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash("Residente actualizado correctamente")
        return redirect(url_for("admin.residentes"))
    return render_template("editar_residente.html", usuario=usuario)


@admin_bp.route("/residentes/bloquear/<id>")
@login_required
@role_required("admin")
def bloquear_residente(id):
    _, col = buscar_residente_por_id(mongo.db, id)
    if col:
        col.update_one({"_id": ObjectId(id)}, {"$set": {"estado": "inactivo"}})
        _invalidar_cache_conteos()

    socketio.emit("actualizar_residentes", to="rol:admin")
    socketio.emit("actualizar_dashboard", to="rol:admin")
    flash("Residente bloqueado correctamente")
    return redirect(url_for("admin.residentes"))


@admin_bp.route("/residentes/eliminar/<id>")
@login_required
@role_required("admin")
def eliminar_residente(id):
    _, col = buscar_residente_por_id(mongo.db, id)
    if col:
        col.delete_one({"_id": ObjectId(id)})
        _invalidar_cache_conteos()

    socketio.emit("actualizar_residentes", to="rol:admin")
    socketio.emit("actualizar_dashboard", to="rol:admin")
    flash("Residente eliminado correctamente.", "success")
    return redirect(url_for("admin.residentes"))


@admin_bp.route("/desbloquear_residente/<id>")
@login_required
@role_required("admin")
def desbloquear_residente(id):
    _, col = buscar_residente_por_id(mongo.db, id)
    if col:
        col.update_one({"_id": ObjectId(id)}, {"$set": {"estado": "activo"}})
        _invalidar_cache_conteos()

    socketio.emit("actualizar_residentes", to="rol:admin")
    flash("Residente desbloqueado correctamente", "success")
    return redirect(url_for("admin.residentes"))


# =========================================
# EXPORTAR RESIDENTES PDF
# =========================================


@admin_bp.route("/residentes/pdf")
@login_required
@role_required("admin")
def exportar_residentes_pdf():
    frac_sel = request.args.get("fraccionamiento", "").strip()
    filtro = {"rol": "residente"}
    if frac_sel:
        filtro["fraccionamiento"] = frac_sel.lower()

    usuarios = list(find_residentes(mongo.db, filtro, sort=[("nombre", 1)]))
    total = contar_residentes(mongo.db, filtro)
    activos = sum(1 for u in usuarios if (u.get("estado") or "").lower() == "activo")
    inactivos = len(usuarios) - activos

    titulo = (
        f"Reporte de Residentes \u2014 {frac_sel.title()}"
        if frac_sel
        else "Reporte General de Residentes"
    )
    subtitulo = frac_sel.title() if frac_sel else "Todos los fraccionamientos"
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(total, "Total residentes", BRAND_ACCENT),
                _kpi_card(activos, "Activos", OK_GREEN),
                _kpi_card(inactivos, "Inactivos", TEXT_MUTED),
                _kpi_card(len(usuarios), "En este reporte", BRAND_DARK),
            ],
            ancho,
        ),
        Spacer(1, 16),
    ]

    if frac_sel:
        data = [["#", "Nombre", "Correo", "Teléfono", "Privada", "Casa", "Estado"]]
        for i, u in enumerate(usuarios, start=1):
            data.append(
                [
                    _c(i, _CELDA_NUM),
                    _c(u.get("nombre", ""), _CELDA_FUERTE),
                    _c(u.get("correo", "")),
                    _c(u.get("telefono", "")),
                    _c(str(u.get("privada", "")).title()),
                    _c(u.get("numero_casa", ""), _CELDA_NUM),
                    _badge_estado(u.get("estado", "")),
                ]
            )
        tabla = _tabla_datos(
            data,
            [30, 150, 235, 95, 90, 50, 75],
            aligns=["CENTER", "LEFT", "LEFT", "CENTER", "LEFT", "CENTER", "CENTER"],
        )
    else:
        data = [
            [
                "#",
                "Nombre",
                "Correo",
                "Teléfono",
                "Fraccionamiento",
                "Privada",
                "Casa",
                "Estado",
            ]
        ]
        for i, u in enumerate(usuarios, start=1):
            data.append(
                [
                    _c(i, _CELDA_NUM),
                    _c(u.get("nombre", ""), _CELDA_FUERTE),
                    _c(u.get("correo", "")),
                    _c(u.get("telefono", "")),
                    _c(str(u.get("fraccionamiento", "")).title()),
                    _c(str(u.get("privada", "")).title()),
                    _c(u.get("numero_casa", ""), _CELDA_NUM),
                    _badge_estado(u.get("estado", "")),
                ]
            )
        tabla = _tabla_datos(
            data,
            [28, 125, 180, 90, 115, 80, 50, 72],
            aligns=[
                "CENTER",
                "LEFT",
                "LEFT",
                "CENTER",
                "LEFT",
                "LEFT",
                "CENTER",
                "CENTER",
            ],
        )

    elementos.append(tabla)
    buffer = BytesIO()
    _construir_reporte_pdf(
        buffer, titulo, subtitulo, elementos, "Reporte de Residentes"
    )
    buffer.seek(0)

    nombre_reporte = (
        f"Reporte_Residentes_{frac_sel.replace(' ', '_')}"
        if frac_sel
        else "Reporte_Residentes_Todos"
    )
    _registrar_historial_reporte(
        f"{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.pdf", "Residentes", "PDF"
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


# =========================================================
# GESTIÓN DE GUARDIAS
# =========================================================


@admin_bp.route("/guardias")
@login_required
@role_required("admin")
def guardias():
    pagina = int(request.args.get("page", 1))
    por_pagina = 10
    total = mongo.db.users.count_documents({"rol": "guardia"})
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    guardias = list(
        mongo.db.users.find({"rol": "guardia"})
        .skip((pagina - 1) * por_pagina)
        .limit(por_pagina)
    )
    return render_template(
        "admin_guardias.html",
        guardias=guardias,
        pagina=pagina,
        total_paginas=total_paginas,
    )


@admin_bp.route("/guardias/registrar", methods=["GET", "POST"])
@login_required
@role_required("admin")
def registrar_guardia():
    if request.method == "POST":
        nombre = request.form["nombre"].strip()
        correo = request.form["correo"].strip()
        telefono = request.form["telefono"].strip()
        turno = request.form["turno"].strip()
        estado = request.form["estado"].strip()
        fraccionamiento = request.form.get("fraccionamiento", "").strip().lower()
        password = request.form.get("password", "").strip()

        disponibles_norm = {f.strip().lower() for f in _fraccionamientos_disponibles()}
        if fraccionamiento not in disponibles_norm:
            flash("Selecciona un fraccionamiento válido para el guardia.", "danger")
            return redirect(url_for("admin.registrar_guardia"))

        if not password:
            password = _generar_password()
        elif len(password) < 8:
            flash("La contraseña debe tener al menos 8 caracteres.", "danger")
            return redirect(url_for("admin.registrar_guardia"))

        if mongo.db.users.find_one({"correo": correo}):
            flash("Ese correo ya se encuentra registrado.", "danger")
            return redirect(url_for("admin.registrar_guardia"))

        mongo.db.users.insert_one(
            {
                "nombre": nombre,
                "correo": correo,
                "password": generate_password_hash(password),
                "telefono": telefono,
                "turno": turno,
                "estado": estado,
                "fraccionamiento": fraccionamiento,
                "rol": "guardia",
                "created_at": datetime.now(),
                "ultimo_acceso": None,
                "intentos_fallidos": 0,
                "bloqueado_hasta": None,
            }
        )

        socketio.emit("actualizar_guardias", to="rol:admin")
        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash(f"Guardia registrado correctamente. Contraseña: {password}", "success")
        return redirect(url_for("admin.guardias"))

    return render_template(
        "registrar_guardia.html",
        fraccionamientos=_fraccionamientos_disponibles(),
    )


@admin_bp.route("/guardias/<id>")
@login_required
@role_required("admin")
def ver_guardia(id):
    guardia = mongo.db.users.find_one({"_id": ObjectId(id)})
    return render_template("ver_guardia.html", guardia=guardia)


@admin_bp.route("/guardias/editar/<id>", methods=["GET", "POST"])
@login_required
@role_required("admin")
def editar_guardia(id):
    guardia = mongo.db.users.find_one({"_id": ObjectId(id)})
    if request.method == "POST":
        mongo.db.users.update_one(
            {"_id": ObjectId(id)},
            {
                "$set": {
                    "nombre": request.form["nombre"],
                    "correo": request.form["correo"],
                    "telefono": request.form["telefono"],
                    "turno": request.form["turno"],
                    "estado": request.form["estado"],
                    "fraccionamiento": request.form.get("fraccionamiento", "")
                    .strip()
                    .lower(),
                }
            },
        )
        socketio.emit("actualizar_guardias", to="rol:admin")
        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash("Guardia actualizado correctamente")
        return redirect(url_for("admin.guardias"))
    return render_template(
        "editar_guardia.html",
        guardia=guardia,
        fraccionamientos=_fraccionamientos_disponibles(),
    )


@admin_bp.route("/guardias/desactivar/<id>")
@login_required
@role_required("admin")
def desactivar_guardia(id):
    mongo.db.users.update_one({"_id": ObjectId(id)}, {"$set": {"estado": "inactivo"}})
    socketio.emit("actualizar_guardias", to="rol:admin")
    socketio.emit("actualizar_dashboard", to="rol:admin")
    flash("Guardia desactivado correctamente")
    return redirect(url_for("admin.guardias"))


@admin_bp.route("/guardias/activar/<id>")
@login_required
@role_required("admin")
def activar_guardia(id):
    mongo.db.users.update_one({"_id": ObjectId(id)}, {"$set": {"estado": "activo"}})
    socketio.emit("actualizar_guardias", to="rol:admin")
    socketio.emit("actualizar_dashboard", to="rol:admin")
    flash("Guardia activado correctamente")
    return redirect(url_for("admin.guardias"))


@admin_bp.route("/guardias/pdf")
@login_required
@role_required("admin")
def exportar_guardias_pdf():
    guardias = list(mongo.db.users.find({"rol": "guardia"}))
    activos = sum(1 for g in guardias if (g.get("estado") or "").lower() == "activo")
    inactivos = len(guardias) - activos
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(len(guardias), "Total guardias", BRAND_ACCENT),
                _kpi_card(activos, "Activos", OK_GREEN),
                _kpi_card(inactivos, "Inactivos", TEXT_MUTED),
            ],
            ancho,
        ),
        Spacer(1, 16),
    ]
    data = [["#", "Nombre", "Correo", "Teléfono", "Turno", "Estado"]]
    for i, g in enumerate(guardias, start=1):
        data.append(
            [
                _c(i, _CELDA_NUM),
                _c(g.get("nombre", ""), _CELDA_FUERTE),
                _c(g.get("correo", "")),
                _c(g.get("telefono", "")),
                _c(str(g.get("turno", "")).title()),
                _badge_estado(g.get("estado", "")),
            ]
        )
    elementos.append(
        _tabla_datos(
            data,
            [35, 170, 235, 120, 90, 95],
            aligns=["CENTER", "LEFT", "LEFT", "CENTER", "CENTER", "CENTER"],
        )
    )

    buffer = BytesIO()
    _construir_reporte_pdf(
        buffer,
        "Reporte de Guardias",
        "Personal de seguridad registrado",
        elementos,
        "Reporte de Guardias",
    )
    buffer.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Guardias_{datetime.now().strftime('%d%m%Y')}.pdf", "Guardias", "PDF"
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Guardias_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


# =========================================================
# HISTORIAL DE ACCESOS
# =========================================================


@admin_bp.route("/accesos")
@login_required
@role_required("admin")
def accesos():
    pagina = int(request.args.get("page", 1))
    por_pagina = 15
    total = mongo.db.access_logs.count_documents({})
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    accesos = list(
        mongo.db.access_logs.find()
        .sort("fecha_hora", -1)
        .skip((pagina - 1) * por_pagina)
        .limit(por_pagina)
    )
    return render_template(
        "admin_accesos.html",
        accesos=accesos,
        pagina=pagina,
        total_paginas=total_paginas,
    )


@admin_bp.route("/accesos/pdf")
@login_required
@role_required("admin")
def exportar_accesos_pdf():
    accesos = list(mongo.db.access_logs.find().sort("fecha_hora", -1))
    rechazados = sum(
        1 for a in accesos if (a.get("resultado") or "").lower() == "rechazado"
    )
    permitidos = len(accesos) - rechazados
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(len(accesos), "Total de accesos", BRAND_ACCENT),
                _kpi_card(permitidos, "Permitidos", OK_GREEN),
                _kpi_card(rechazados, "Rechazados", WARN_RED),
            ],
            ancho,
        ),
        Spacer(1, 16),
    ]
    data = [["#", "Guardia", "Acción", "Resultado", "Fecha"]]
    for i, a in enumerate(accesos, start=1):
        fecha = a.get("fecha_hora")
        fecha = fecha.strftime("%d/%m/%Y %I:%M %p") if fecha else ""
        data.append(
            [
                _c(i, _CELDA_NUM),
                _c(a.get("guardia_nombre", ""), _CELDA_FUERTE),
                _c(a.get("accion", "")),
                _badge_estado(a.get("resultado", "")),
                _c(fecha),
            ]
        )
    elementos.append(
        _tabla_datos(
            data,
            [35, 180, 200, 145, 165],
            aligns=["CENTER", "LEFT", "LEFT", "CENTER", "CENTER"],
        )
    )

    buffer = BytesIO()
    _construir_reporte_pdf(
        buffer,
        "Reporte de Accesos",
        "Registro de entradas y validaciones de QR",
        elementos,
        "Reporte de Accesos",
    )
    buffer.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Accesos_{datetime.now().strftime('%d%m%Y')}.pdf", "Accesos", "PDF"
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Accesos_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


@admin_bp.route("/accesos/excel")
@login_required
@role_required("admin")
def exportar_accesos_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    accesos = list(mongo.db.access_logs.find().sort("fecha_hora", -1))
    data = []
    for a in accesos:
        fecha = a.get("fecha_hora")
        data.append(
            {
                "Guardia": a.get("guardia_nombre", ""),
                "Acción": a.get("accion", ""),
                "Resultado": a.get("resultado", ""),
                "Fecha": fecha.strftime("%d/%m/%Y %I:%M %p") if fecha else "",
            }
        )
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Accesos", index=False)
        ws = writer.sheets["Accesos"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        for col, w in {"A": 28, "B": 35, "C": 40, "D": 25}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Accesos_{datetime.now().strftime('%d%m%Y')}.xlsx", "Accesos", "Excel"
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Accesos_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# INCIDENCIAS
# =========================================================


@admin_bp.route("/incidencias")
@login_required
@role_required("admin")
def incidencias():
    pagina = int(request.args.get("page", 1))
    por_pagina = 15
    total = mongo.db.incidencias.count_documents({})
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    incidencias = list(
        mongo.db.incidencias.find()
        .sort("fecha_hora", -1)
        .skip((pagina - 1) * por_pagina)
        .limit(por_pagina)
    )
    return render_template(
        "admin_incidencias.html",
        incidencias=incidencias,
        pagina=pagina,
        total_paginas=total_paginas,
    )


@admin_bp.route("/incidencias/pdf")
@login_required
@role_required("admin")
def exportar_incidencias_pdf():
    incidencias = list(mongo.db.incidencias.find().sort("fecha_hora", -1))
    _resueltos = {"resuelta", "resuelto", "cerrada", "cerrado"}
    resueltas = sum(
        1 for x in incidencias if (x.get("estado") or "").lower() in _resueltos
    )
    abiertas = len(incidencias) - resueltas
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(len(incidencias), "Total incidencias", BRAND_ACCENT),
                _kpi_card(abiertas, "Abiertas / pendientes", AMBER),
                _kpi_card(resueltas, "Resueltas", OK_GREEN),
            ],
            ancho,
        ),
        Spacer(1, 16),
    ]
    data = [["#", "Tipo", "Guardia", "Descripción", "Estado", "Fecha"]]
    for i, inc in enumerate(incidencias, start=1):
        fecha = inc.get("fecha_hora")
        fecha = fecha.strftime("%d/%m/%Y %I:%M %p") if fecha else ""
        data.append(
            [
                _c(i, _CELDA_NUM),
                _c(str(inc.get("tipo_incidencia", "")).title(), _CELDA_FUERTE),
                _c(inc.get("guardia_nombre", "")),
                _c(inc.get("descripcion", "Sin descripción")),
                _badge_estado(inc.get("estado", "")),
                _c(fecha),
            ]
        )
    elementos.append(
        _tabla_datos(
            data,
            [28, 95, 105, 300, 85, 110],
            aligns=["CENTER", "LEFT", "LEFT", "LEFT", "CENTER", "CENTER"],
        )
    )

    buffer = BytesIO()
    _construir_reporte_pdf(
        buffer,
        "Reporte de Incidencias",
        "Eventos reportados por el personal",
        elementos,
        "Reporte de Incidencias",
    )
    buffer.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Incidencias_{datetime.now().strftime('%d%m%Y')}.pdf",
        "Incidencias",
        "PDF",
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Incidencias_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


@admin_bp.route("/incidencias/excel")
@login_required
@role_required("admin")
def exportar_incidencias_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    incidencias = list(mongo.db.incidencias.find().sort("fecha_hora", -1))
    data = []
    for inc in incidencias:
        fecha = inc.get("fecha_hora")
        data.append(
            {
                "Tipo Incidencia": inc.get("tipo_incidencia", ""),
                "Guardia": inc.get("guardia_nombre", ""),
                "Descripción": inc.get("descripcion", ""),
                "Estado": inc.get("estado", ""),
                "Fecha y Hora": fecha.strftime("%d/%m/%Y %I:%M %p") if fecha else "",
            }
        )
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Incidencias", index=False)
        ws = writer.sheets["Incidencias"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top", horizontal="center", wrap_text=True
                )
                cell.border = border
        for col, w in {"A": 25, "B": 25, "C": 70, "D": 20, "E": 25}.items():
            ws.column_dimensions[col].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    output.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Incidencias_{datetime.now().strftime('%d%m%Y')}.xlsx",
        "Incidencias",
        "Excel",
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Incidencias_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# REPORTES
# =========================================================


@admin_bp.route("/reportes")
@login_required
@role_required("admin")
def reportes():
    return render_template("admin_reportes.html")


# =========================================================
# EXPORTAR VISITAS PDF / EXCEL
# =========================================================


@admin_bp.route("/visitas/pdf")
@login_required
@role_required("admin")
def exportar_visitas_pdf():
    from datetime import datetime as _dt

    MESES_ES = [
        "",
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ]

    frac_sel = request.args.get("fraccionamiento", "").strip().lower()
    modo = request.args.get("modo", "dia")
    dia_sel = request.args.get("dia", "").strip()
    semana_sel = request.args.get("semana", "").strip()
    mes_sel = request.args.get("mes", "").strip()
    fecha_inicio_tabla = request.args.get("fecha_inicio_tabla", "").strip()
    fecha_fin_tabla = request.args.get("fecha_fin_tabla", "").strip()
    busqueda = request.args.get("busqueda", "").strip()
    hoy = _ahora_local()

    if modo == "semana":
        if not semana_sel:
            iso = hoy.isocalendar()
            semana_sel = f"{iso[0]}-W{iso[1]:02d}"
        anio, sem = semana_sel.split("-W")
        ini = _dt.strptime(f"{anio}-W{int(sem):02d}-1", "%G-W%V-%u")
        fin = ini + timedelta(days=6)
        periodo_label = (
            f"Semana del {ini.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
        )
    elif modo == "mes":
        if not mes_sel:
            mes_sel = hoy.strftime("%Y-%m")
        anio, mes = map(int, mes_sel.split("-"))
        ini = datetime(anio, mes, 1)
        fin = (
            datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        ) - timedelta(days=1)
        periodo_label = f"{MESES_ES[mes]} {anio}"
    else:
        if not dia_sel:
            dia_sel = hoy.strftime("%Y-%m-%d")
        ini = _dt.strptime(dia_sel, "%Y-%m-%d")
        fin = ini
        periodo_label = ini.strftime("%d/%m/%Y")

    inicio_str = ini.strftime("%Y-%m-%d")
    fin_str = fin.strftime("%Y-%m-%d")
    if fecha_inicio_tabla or fecha_fin_tabla:
        inicio_str = fecha_inicio_tabla or inicio_str
        fin_str = fecha_fin_tabla or fin_str
        if inicio_str > fin_str:
            inicio_str, fin_str = fin_str, inicio_str
        periodo_label = f"Del {inicio_str} al {fin_str}"

    filtro = {"fecha_visita": {"$gte": inicio_str, "$lte": fin_str}}
    if frac_sel:
        filtro["fraccionamiento"] = frac_sel
    if busqueda:
        import re as _re

        rx = {"$regex": _re.escape(busqueda), "$options": "i"}
        filtro["$or"] = [
            {"nombre_visitante": rx},
            {"residente_nombre": rx},
            {"vehiculo.placa": rx},
        ]

    visitas = list(
        find_visitas(
            mongo.db,
            filtro,
            sort=[("fecha_visita", 1), ("hora_inicio", 1)],
            frac=frac_sel or None,
        )
    )

    estado_map = {
        "activo": "Activo",
        "dentro": "Dentro",
        "salida_registrada": "Finalizada",
        "cancelado": "Cancelado",
        "vencido": "Vencido",
    }

    def _cuenta(*estados):
        return sum(1 for v in visitas if v.get("estado") in estados)

    titulo = "Reporte de Visitas"
    subtitulo = (
        f"{frac_sel.title()} \u00b7 {periodo_label}"
        if frac_sel
        else f"Todos los fraccionamientos \u00b7 {periodo_label}"
    )
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(len(visitas), "Total visitas", BRAND_ACCENT),
                _kpi_card(_cuenta("dentro"), "Dentro", CYAN),
                _kpi_card(_cuenta("activo"), "Activas", OK_GREEN),
                _kpi_card(_cuenta("salida_registrada"), "Finalizadas", TEXT_MUTED),
            ],
            ancho,
        ),
        Spacer(1, 16),
    ]
    data = [["#", "Visitante", "Residente", "Placas", "Estado"]]
    for i, v in enumerate(visitas, start=1):
        vehiculo = v.get("vehiculo", {})
        placa = vehiculo.get("placa", "") if isinstance(vehiculo, dict) else ""
        data.append(
            [
                _c(i, _CELDA_NUM),
                _c(v.get("nombre_visitante", ""), _CELDA_FUERTE),
                _c(v.get("residente_nombre", "")),
                _c(str(placa).upper(), _CELDA_NUM),
                _badge_estado(
                    v.get("estado", ""),
                    estado_map.get(v.get("estado", ""), v.get("estado", "")),
                ),
            ]
        )
    elementos.append(
        _tabla_datos(
            data,
            [35, 240, 230, 110, 95],
            aligns=["CENTER", "LEFT", "LEFT", "CENTER", "CENTER"],
        )
    )

    buffer = BytesIO()
    _construir_reporte_pdf(buffer, titulo, subtitulo, elementos, "Reporte de Visitas")
    buffer.seek(0)
    nombre_reporte = (
        f"Reporte_Visitas_{frac_sel.replace(' ', '_')}"
        if frac_sel
        else "Reporte_Visitas_Todos"
    )
    _registrar_historial_reporte(
        f"{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.pdf", "Visitas", "PDF"
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


@admin_bp.route("/visitas/excel")
@login_required
@role_required("admin")
def exportar_visitas_excel():
    from datetime import datetime as _dt
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    frac_sel = request.args.get("fraccionamiento", "").strip().lower()
    modo = request.args.get("modo", "dia")
    dia_sel = request.args.get("dia", "").strip()
    semana_sel = request.args.get("semana", "").strip()
    mes_sel = request.args.get("mes", "").strip()
    fecha_inicio_tabla = request.args.get("fecha_inicio_tabla", "").strip()
    fecha_fin_tabla = request.args.get("fecha_fin_tabla", "").strip()
    busqueda = request.args.get("busqueda", "").strip()
    hoy = _ahora_local()

    if modo == "semana":
        if not semana_sel:
            iso = hoy.isocalendar()
            semana_sel = f"{iso[0]}-W{iso[1]:02d}"
        anio, sem = semana_sel.split("-W")
        ini = _dt.strptime(f"{anio}-W{int(sem):02d}-1", "%G-W%V-%u")
        fin = ini + timedelta(days=6)
    elif modo == "mes":
        if not mes_sel:
            mes_sel = hoy.strftime("%Y-%m")
        anio, mes = map(int, mes_sel.split("-"))
        ini = datetime(anio, mes, 1)
        fin = (
            datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        ) - timedelta(days=1)
    else:
        if not dia_sel:
            dia_sel = hoy.strftime("%Y-%m-%d")
        ini = _dt.strptime(dia_sel, "%Y-%m-%d")
        fin = ini

    inicio_str = ini.strftime("%Y-%m-%d")
    fin_str = fin.strftime("%Y-%m-%d")
    if fecha_inicio_tabla or fecha_fin_tabla:
        inicio_str = fecha_inicio_tabla or inicio_str
        fin_str = fecha_fin_tabla or fin_str
        if inicio_str > fin_str:
            inicio_str, fin_str = fin_str, inicio_str

    filtro = {"fecha_visita": {"$gte": inicio_str, "$lte": fin_str}}
    if frac_sel:
        filtro["fraccionamiento"] = frac_sel
    if busqueda:
        import re as _re

        rx = {"$regex": _re.escape(busqueda), "$options": "i"}
        filtro["$or"] = [
            {"nombre_visitante": rx},
            {"residente_nombre": rx},
            {"vehiculo.placa": rx},
        ]

    visitas = find_visitas(
        mongo.db,
        filtro,
        sort=[("fecha_visita", 1), ("hora_inicio", 1)],
        frac=frac_sel or None,
    )
    estado_map = {
        "activo": "Activo",
        "dentro": "Dentro",
        "salida_registrada": "Finalizada",
        "cancelado": "Cancelado",
        "vencido": "Vencido",
    }
    data = []
    for v in visitas:
        vehiculo = v.get("vehiculo", {})
        placa = vehiculo.get("placa", "") if isinstance(vehiculo, dict) else ""
        data.append(
            {
                "Visitante": v.get("nombre_visitante", ""),
                "Residente": v.get("residente_nombre", ""),
                "Fraccionamiento": str(v.get("fraccionamiento", "")).title(),
                "Fecha": v.get("fecha_visita", ""),
                "Placas": placa,
                "Estado": estado_map.get(v.get("estado", ""), v.get("estado", "")),
            }
        )

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Visitas", index=False)
        ws = writer.sheets["Visitas"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    nombre_reporte = (
        f"Reporte_Visitas_{frac_sel.replace(' ', '_')}"
        if frac_sel
        else "Reporte_Visitas_Todos"
    )
    _registrar_historial_reporte(
        f"{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.xlsx", "Visitas", "Excel"
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_{nombre_reporte}_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/residentes/excel")
@login_required
@role_required("admin")
def exportar_residentes_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    residentes = find_residentes(mongo.db, {"rol": "residente"})
    data = [
        {
            "Nombre": r.get("nombre", ""),
            "Correo": r.get("correo", ""),
            "Telefono": r.get("telefono", ""),
            "Privada": r.get("privada", ""),
            "Casa": r.get("numero_casa", ""),
            "Estado": r.get("estado", ""),
        }
        for r in residentes
    ]

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Residentes", index=False)
        ws = writer.sheets["Residentes"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
        for col, w in {"A": 35, "B": 35, "C": 22, "D": 22, "E": 15, "F": 18}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Residentes_{datetime.now().strftime('%d%m%Y')}.xlsx",
        "Residentes",
        "Excel",
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Residentes_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/guardias/excel")
@login_required
@role_required("admin")
def exportar_guardias_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    guardias = list(mongo.db.users.find({"rol": "guardia"}))
    data = [
        {
            "Nombre": g.get("nombre", ""),
            "Correo": g.get("correo", ""),
            "Telefono": g.get("telefono", ""),
            "Turno": g.get("turno", ""),
            "Estado": g.get("estado", ""),
        }
        for g in guardias
    ]

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Guardias", index=False)
        ws = writer.sheets["Guardias"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
        for col, w in {"A": 35, "B": 35, "C": 22, "D": 18, "E": 18}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    _registrar_historial_reporte(
        f"Reporte_Guardias_{datetime.now().strftime('%d%m%Y')}.xlsx",
        "Guardias",
        "Excel",
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_Guardias_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# EXPORTAR SISTEMA PDF / EXCEL
# =========================================================


@admin_bp.route("/sistema/pdf")
@login_required
@role_required("admin")
def exportar_sistema_pdf():
    total_residentes = contar_residentes(mongo.db, {"rol": "residente"})
    total_guardias = mongo.db.users.count_documents({"rol": "guardia"})
    total_visitas = contar_visitas(mongo.db)
    total_accesos = mongo.db.access_logs.count_documents({})
    total_incidencias = mongo.db.incidencias.count_documents({})
    ancho = _ancho_util()

    elementos = [
        _fila_kpis(
            [
                _kpi_card(total_residentes, "Residentes", BRAND_ACCENT),
                _kpi_card(total_guardias, "Guardias", OK_GREEN),
                _kpi_card(total_visitas, "Visitas", CYAN),
                _kpi_card(total_accesos, "Accesos", AMBER),
                _kpi_card(total_incidencias, "Incidencias", WARN_RED),
            ],
            ancho,
        ),
        Spacer(1, 22),
    ]
    data = [
        ["Módulo", "Total de registros"],
        [_c("Residentes", _CELDA_FUERTE), _c(total_residentes, _CELDA_NUM)],
        [_c("Guardias", _CELDA_FUERTE), _c(total_guardias, _CELDA_NUM)],
        [_c("Visitas", _CELDA_FUERTE), _c(total_visitas, _CELDA_NUM)],
        [_c("Accesos", _CELDA_FUERTE), _c(total_accesos, _CELDA_NUM)],
        [_c("Incidencias", _CELDA_FUERTE), _c(total_incidencias, _CELDA_NUM)],
    ]
    elementos.append(
        _tabla_datos(data, [ancho * 0.62, ancho * 0.38], aligns=["LEFT", "CENTER"])
    )

    buffer = BytesIO()
    _construir_reporte_pdf(
        buffer,
        "Reporte General del Sistema",
        "Resumen consolidado de todos los módulos",
        elementos,
        "Reporte General",
    )
    buffer.seek(0)
    _registrar_historial_reporte(
        f"Reporte_General_{datetime.now().strftime('%d%m%Y')}.pdf", "General", "PDF"
    )
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_General_{datetime.now().strftime('%d%m%Y')}.pdf",
        mimetype="application/pdf",
    )


@admin_bp.route("/sistema/excel")
@login_required
@role_required("admin")
def exportar_sistema_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = [
        {
            "Módulo": "Residentes",
            "Total": contar_residentes(mongo.db, {"rol": "residente"}),
        },
        {
            "Módulo": "Guardias",
            "Total": mongo.db.users.count_documents({"rol": "guardia"}),
        },
        {"Módulo": "Visitas", "Total": contar_visitas(mongo.db)},
        {"Módulo": "Accesos", "Total": mongo.db.access_logs.count_documents({})},
        {"Módulo": "Incidencias", "Total": mongo.db.incidencias.count_documents({})},
    ]
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sistema", index=False)
        ws = writer.sheets["Sistema"]
        hf = PatternFill("solid", fgColor="0F172A")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
        for col, w in {"A": 35, "B": 20}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    _registrar_historial_reporte(
        f"Reporte_General_{datetime.now().strftime('%d%m%Y')}.xlsx", "General", "Excel"
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Acceso_QR_Reporte_General_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# HISTORIAL DE REPORTES
# =========================================================


@admin_bp.route("/historial-reportes")
@login_required
@role_required("admin")
def historial_reportes():
    pagina = int(request.args.get("page", 1))
    por_pagina = 15
    total = mongo.db.reportes.count_documents({})
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    reportes = list(
        mongo.db.reportes.find()
        .sort("fecha", -1)
        .skip((pagina - 1) * por_pagina)
        .limit(por_pagina)
    )
    for r in reportes:
        if isinstance(r.get("fecha"), datetime):
            r["fecha"] = r["fecha"].strftime("%d/%m/%Y %I:%M %p")
    return render_template(
        "admin_historial_reportes.html",
        reportes=reportes,
        pagina=pagina,
        total_paginas=total_paginas,
    )


# =========================================================
# CONFIGURACIÓN POR FRACCIONAMIENTO
# =========================================================


def _slug_fraccionamiento(nombre):
    import re, unicodedata

    s = unicodedata.normalize("NFKD", nombre or "").encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "_", s.lower().strip()).strip("_")
    return s


def _formato_frac(nombre):
    """Mayúsculas iniciales como en la UI: 'villas del bosque ii' -> 'Villas del Bosque II'."""
    s = (nombre or "").title()
    s = s.replace(" Del ", " del ")
    s = s.replace(" Ii", " II").replace(" Iii", " III")
    return s


def _contar_residentes_frac(nombre):
    col = f"residentes_{_slug_fraccionamiento(nombre)}"
    if col in mongo.db.list_collection_names():
        return mongo.db[col].count_documents({"rol": "residente"})
    return 0


def _config_fraccionamiento(frac):
    slug = _slug_fraccionamiento(frac)
    base = {
        "fraccionamiento": slug,
        "nombre": frac,
        "direccion": "",
        "telefono": "",
        "correo": "",
        "privadas": 1,
        "duracion_qr": "3 horas",
        "multiples_accesos": "No",
        "validacion_vehiculo": "Obligatoria",
        "foto_visitante": "Sí",
        "registro_incidencias": "Habilitado",
        "control_sesiones": True,
        "bitacora": True,
        "respaldos": True,
        "camaras": False,
        "lectura_placas": False,
        "notificaciones": True,
        "actualizado": None,
    }
    doc = mongo.db.config_fraccionamientos.find_one({"fraccionamiento": slug}) or {}
    for k, v in doc.items():
        if k != "_id":
            base[k] = v
    act = base.get("actualizado")
    base["actualizado_str"] = (
        act.strftime("%d/%m/%Y %I:%M %p") if isinstance(act, datetime) else ""
    )
    return base


@admin_bp.route("/fraccionamientos/crear", methods=["POST"])
@login_required
@role_required("admin")
def crear_fraccionamiento():
    nombre = request.form.get("nombre", "").strip()
    if len(nombre) < 3:
        flash("Escribe un nombre válido (mínimo 3 caracteres).", "danger")
        return redirect(url_for("admin.configuracion"))

    slug = _slug_fraccionamiento(nombre)
    if not slug:
        flash("El nombre no genera un identificador válido.", "danger")
        return redirect(url_for("admin.configuracion"))

    slugs_existentes = {
        _slug_fraccionamiento(f) for f in _fraccionamientos_disponibles()
    }
    if slug in slugs_existentes:
        flash(f"El fraccionamiento «{_formato_frac(nombre)}» ya existe.", "danger")
        return redirect(url_for("admin.configuracion", fraccionamiento=nombre))

    existentes = set(mongo.db.list_collection_names())
    for prefijo in ("residentes", "visitas"):
        col = f"{prefijo}_{slug}"
        if col not in existentes:
            mongo.db.create_collection(col)

    mongo.db.fraccionamientos.insert_one(
        {"nombre": nombre, "slug": slug, "created_at": datetime.now()}
    )
    mongo.db.config_fraccionamientos.update_one(
        {"fraccionamiento": slug},
        {
            "$setOnInsert": {
                "fraccionamiento": slug,
                "nombre": nombre,
                "privadas": 1,
                "actualizado": None,
            }
        },
        upsert=True,
    )

    _invalidar_cache_fracs()
    _invalidar_cache_conteos()

    socketio.emit("actualizar_dashboard", to="rol:admin")
    flash(f"Fraccionamiento «{_formato_frac(nombre)}» creado correctamente.", "success")
    return redirect(url_for("admin.configuracion", fraccionamiento=nombre))


@admin_bp.route("/eliminar_fraccionamiento", methods=["POST"])
@login_required
@role_required("admin")
def eliminar_fraccionamiento():
    nombre = request.form.get("fraccionamiento", "").strip()
    if not nombre:
        flash("Fraccionamiento inválido.", "danger")
        return redirect(url_for("admin.configuracion"))

    disponibles = _fraccionamientos_disponibles()
    if len(disponibles) <= 1:
        flash("No puedes eliminar el único fraccionamiento del sistema.", "danger")
        return redirect(url_for("admin.configuracion"))

    slug = _slug_fraccionamiento(nombre)
    try:
        mongo.db.drop_collection(f"residentes_{slug}")
        mongo.db.drop_collection(f"visitas_{slug}")
        mongo.db.config_fraccionamientos.delete_one({"fraccionamiento": slug})
        mongo.db.fraccionamientos.delete_one({"slug": slug})

        _invalidar_cache_fracs()
        _invalidar_cache_conteos()

        socketio.emit("actualizar_dashboard", to="rol:admin")
        flash(
            f"Fraccionamiento '{_formato_frac(nombre)}' eliminado correctamente.",
            "success",
        )
    except Exception as e:
        flash(f"Error al eliminar fraccionamiento: {str(e)}", "danger")

    return redirect(url_for("admin.configuracion"))


@admin_bp.route("/configuracion", methods=["GET", "POST"])
@login_required
@role_required("admin")
def configuracion():
    disponibles = _fraccionamientos_disponibles()
    frac_sel = request.args.get("fraccionamiento", "").strip()
    if request.method == "POST":
        frac_sel = request.form.get("fraccionamiento", frac_sel).strip()
    if frac_sel not in disponibles:
        frac_sel = disponibles[0] if disponibles else ""

    slug = _slug_fraccionamiento(frac_sel)

    if request.method == "POST":
        seccion = request.form.get("seccion", "")
        datos = {}
        if seccion == "general":
            try:
                privadas = int(request.form.get("privadas", 1) or 1)
            except ValueError:
                privadas = 1
            datos = {
                "nombre": request.form.get("nombre", "").strip() or frac_sel,
                "direccion": request.form.get("direccion", "").strip(),
                "telefono": request.form.get("telefono", "").strip(),
                "correo": request.form.get("correo", "").strip().lower(),
                "privadas": max(1, min(99, privadas)),
            }
        elif seccion == "accesos":
            for c in (
                "duracion_qr",
                "multiples_accesos",
                "validacion_vehiculo",
                "foto_visitante",
                "registro_incidencias",
            ):
                datos[c] = request.form.get(c, "")
        elif seccion == "seguridad":
            for c in ("control_sesiones", "bitacora", "respaldos"):
                datos[c] = request.form.get(c) == "on"
        elif seccion == "avanzadas":
            for c in ("camaras", "lectura_placas", "notificaciones"):
                datos[c] = request.form.get(c) == "on"

        if datos:
            datos["fraccionamiento"] = slug
            datos["actualizado"] = datetime.now()
            mongo.db.config_fraccionamientos.update_one(
                {"fraccionamiento": slug}, {"$set": datos}, upsert=True
            )
            flash(
                f"Configuración de {_formato_frac(frac_sel)} guardada correctamente.",
                "success",
            )

        return redirect(url_for("admin.configuracion", fraccionamiento=frac_sel))

    config = _config_fraccionamiento(frac_sel)

    _, total_g = _conteos_residentes_cacheados(disponibles)
    conteos_cache, _ = _conteos_residentes_cacheados(disponibles)
    conteos = conteos_cache

    return render_template(
        "admin_configuracion.html",
        fraccionamientos=disponibles,
        frac_sel=frac_sel,
        config=config,
        conteos=conteos,
    )
